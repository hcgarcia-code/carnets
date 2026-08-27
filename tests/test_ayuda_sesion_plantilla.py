"""Tres huecos que dejan al sindicato sin salida en mitad de su tarea.

El boton de ayuda de la portada abria el programa de correo del ordenador. En
un ordenador compartido de local sindical no hay ninguno configurado, asi que
el boton no hacia nada visible: el unico camino de ayuda de la web terminaba en
una pantalla en blanco.

Una vez dentro no habia forma de cerrar sesion. En equipos compartidos eso
significa que la sesion del sindicato anterior sigue abierta para quien se
siente despues, que es exactamente el escenario que el cifrado en reposo y la
auditoria de accesos estan para evitar.

Y la pagina de carga pedia "usa la plantilla oficial" sin dar forma de
conseguirla. La plantilla existia, pero en el escritorio de una persona.
"""

import pytest
from django.contrib.auth.models import User
from django.contrib.staticfiles import finders

from gestion.models import PerfilSindicato
from gestion.views import COLUMNAS_REQUERIDAS

PLANTILLA = "documentos/plantilla_afiliacion_oficial.xlsx"


@pytest.fixture
def sindicato(client):
    usuario = User.objects.create_user("sov_prueba", "s@ejemplo.es", "clave-larga-123")  # noqa: S106
    PerfilSindicato.objects.create(usuario=usuario, acuerdo_aceptado=True)
    client.force_login(usuario)
    return usuario


# --- La pagina de ayuda ---

@pytest.mark.django_db
def test_la_pagina_de_ayuda_responde(client):
    assert client.get("/ayuda/").status_code == 200


@pytest.mark.django_db
def test_la_ayuda_se_ve_sin_tener_cuenta(client):
    """Quien mas la necesita es justo quien no consigue entrar."""
    respuesta = client.get("/ayuda/")

    assert respuesta.status_code == 200
    assert "login" not in respuesta.headers.get("Location", "")


@pytest.mark.django_db
def test_la_ayuda_da_las_dos_formas_de_contacto(client):
    contenido = client.get("/ayuda/").content.decode()

    assert "soporte-carnets@cgt.org.es" in contenido
    assert "918 055 047" in contenido or "918055047" in contenido


@pytest.mark.django_db
def test_el_boton_de_ayuda_de_la_portada_lleva_a_la_ayuda(client):
    """El fallo original: el boton era un mailto:. Se comprueba sobre el HTML
    servido, que es lo que recibe el navegador."""
    contenido = client.get("/").content.decode()

    assert 'href="/ayuda/"' in contenido, "la portada no enlaza la pagina de ayuda"


@pytest.mark.django_db
def test_ningun_boton_de_la_portada_abre_el_programa_de_correo(client):
    """Un mailto: en un boton de accion depende de que el ordenador tenga
    cliente de correo configurado. En un local sindical no lo tiene."""
    contenido = client.get("/").content.decode()

    for trozo in contenido.split("<a ")[1:]:
        etiqueta = trozo.split(">")[0]
        if "boton" in etiqueta:
            assert "mailto:" not in etiqueta, f"boton con mailto: {etiqueta[:120]}"


@pytest.mark.django_db
def test_la_ayuda_no_dice_quien_usa_el_sistema(client):
    """Es publica, como la portada, y le aplica el mismo criterio."""
    usuario = User.objects.create_user("sov_secreto", "b@ejemplo.es", "clave-larga-123")  # noqa: S106
    PerfilSindicato.objects.create(
        usuario=usuario, nombre_identificativo="SOV Madrid", persona_contacto="Ana Ejemplo",
    )

    contenido = client.get("/ayuda/").content.decode()

    assert "SOV Madrid" not in contenido
    assert "Ana Ejemplo" not in contenido


@pytest.mark.django_db
def test_la_ayuda_usa_la_hoja_comun(client):
    assert "gestion/carnets.css" in client.get("/ayuda/").content.decode()


# --- Cerrar sesion ---

@pytest.mark.django_db
def test_se_puede_cerrar_sesion(client, sindicato):
    respuesta = client.post("/salir/")

    assert respuesta.status_code == 302
    assert "_auth_user_id" not in client.session, "la sesion sigue abierta"


@pytest.mark.django_db
def test_al_cerrar_sesion_vuelve_a_la_portada(client, sindicato):
    """Dejarlo en el login invita a volver a entrar; la portada es la salida."""
    assert client.post("/salir/")["Location"] == "/"


@pytest.mark.django_db
def test_despues_de_salir_la_zona_privada_pide_entrar(client, sindicato):
    """La comprobacion que de verdad importa: que la sesion se cerrase, no que
    la pagina dijera que se cerro."""
    client.post("/salir/")

    respuesta = client.get("/subir-datos/")

    assert respuesta.status_code == 302
    assert "/login/" in respuesta["Location"]


@pytest.mark.django_db
def test_un_enlace_ajeno_no_puede_cerrarte_la_sesion(client, sindicato):
    """Django 5 solo admite POST justo por esto: con GET, bastaba una imagen
    con src=/salir/ en cualquier pagina para echar al sindicato de la suya."""
    respuesta = client.get("/salir/")

    assert respuesta.status_code == 405
    assert "_auth_user_id" in client.session, "un GET ha cerrado la sesion"


@pytest.mark.django_db
def test_la_barra_ofrece_salir_cuando_hay_sesion(client, sindicato):
    contenido = client.get("/subir-datos/").content.decode()

    assert 'action="/salir/"' in contenido, "no hay forma de cerrar sesion en la barra"


@pytest.mark.django_db
def test_la_barra_no_ofrece_salir_sin_sesion(client):
    assert "/salir/" not in client.get("/").content.decode()


# --- La plantilla oficial ---

def test_la_plantilla_esta_donde_django_la_busca():
    assert finders.find(PLANTILLA) is not None, f"no se encuentra {PLANTILLA}"


@pytest.mark.django_db
def test_la_pagina_de_carga_ofrece_la_plantilla(client, sindicato):
    """Pedia usar la plantilla oficial sin decir de donde sacarla."""
    contenido = client.get("/subir-datos/").content.decode()

    assert PLANTILLA in contenido, "la pagina de carga no enlaza la plantilla"


def test_la_plantilla_trae_las_columnas_que_el_sistema_exige():
    """La prueba que justifica el archivo. Si alguien toca COLUMNAS_REQUERIDAS
    y no la plantilla, los sindicatos se descargan un archivo que el validador
    rechaza, y el error que ven ("faltan columnas obligatorias") les senala a
    ellos un fallo que no es suyo."""
    import openpyxl

    libro = openpyxl.load_workbook(finders.find(PLANTILLA))
    cabecera = [c.value for c in libro[libro.sheetnames[0]][1]]

    assert cabecera == COLUMNAS_REQUERIDAS, (
        f"la plantilla trae {cabecera} y el sistema espera {COLUMNAS_REQUERIDAS}"
    )


def test_la_plantilla_va_vacia_de_personas():
    """Se sirve desde /static/, o sea a cualquiera que sepa la direccion. Una
    fila de ejemplo con un nombre real seria una filtracion permanente."""
    import openpyxl

    libro = openpyxl.load_workbook(finders.find(PLANTILLA))
    altas = libro[libro.sheetnames[0]]

    assert altas.max_row == 1, f"la hoja de altas trae {altas.max_row - 1} filas de datos"
